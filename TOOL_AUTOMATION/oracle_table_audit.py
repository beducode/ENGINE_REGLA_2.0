import json
import csv
import traceback
from pathlib import Path
from datetime import datetime

import oracledb
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# =========================================================
# PATHS
# =========================================================
BASE_DIR = Path(__file__).resolve().parent
CONFIG_PATH = BASE_DIR / "config.json"
OUTPUT_DIR = BASE_DIR / "audit_output"

# =========================================================
# CONFIG
# =========================================================
def load_raw_config():
    if not CONFIG_PATH.exists():
        raise FileNotFoundError(f"config.json tidak ditemukan: {CONFIG_PATH}")

    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        return json.load(f)

def normalize_config(raw_cfg: dict) -> dict:
    cfg = dict(raw_cfg)

    user = (
        cfg.get("user")
        or cfg.get("username")
        or cfg.get("db_user")
    )

    password = (
        cfg.get("password")
        or cfg.get("db_password")
    )

    dsn = (
        cfg.get("dsn")
        or cfg.get("db_dsn")
    )

    if not dsn:
        host = cfg.get("host")
        port = cfg.get("port", 1521)
        service_name = cfg.get("service_name") or cfg.get("sid")

        if host and service_name:
            dsn = f"{host}:{port}/{service_name}"

    final_cfg = {
        "user": user,
        "password": password,
        "dsn": dsn
    }

    validate_config(final_cfg)
    return final_cfg

def validate_config(cfg: dict):
    missing = []

    if not cfg.get("user"):
        missing.append("user (atau username / db_user)")
    if not cfg.get("password"):
        missing.append("password (atau db_password)")
    if not cfg.get("dsn"):
        missing.append("dsn (atau db_dsn / host+port+service_name)")

    if missing:
        raise ValueError(
            "Config tidak lengkap. Field berikut wajib ada:\n- "
            + "\n- ".join(missing)
        )

def load_config():
    raw = load_raw_config()
    return normalize_config(raw)

# =========================================================
# DB
# =========================================================
def connect_oracle(cfg: dict):
    return oracledb.connect(
        user=cfg["user"],
        password=cfg["password"],
        dsn=cfg["dsn"]
    )

# =========================================================
# HELPERS
# =========================================================
def quote_ident(name: str) -> str:
    return f'"{name.upper()}"'

def now_str():
    return datetime.now().strftime("%Y%m%d_%H%M%S")

def log(msg: str):
    print(msg)

# =========================================================
# ORACLE METADATA
# =========================================================
def get_all_tables(cursor):
    cursor.execute("""
        SELECT TABLE_NAME
        FROM USER_TABLES
        ORDER BY TABLE_NAME
    """)
    return [row[0] for row in cursor.fetchall()]

def get_table_stats(cursor):
    cursor.execute("""
        SELECT TABLE_NAME, NUM_ROWS, LAST_ANALYZED
        FROM USER_TABLES
        ORDER BY TABLE_NAME
    """)
    stats = {}
    for row in cursor.fetchall():
        table_name, num_rows, last_analyzed = row
        stats[table_name] = {
            "num_rows": num_rows,
            "last_analyzed": str(last_analyzed) if last_analyzed else None
        }
    return stats

def count_table_rows_exact(cursor, table_name: str):
    sql = f'SELECT COUNT(*) FROM {quote_ident(table_name)}'
    cursor.execute(sql)
    return cursor.fetchone()[0]

# =========================================================
# AUDIT
# =========================================================
def audit_tables(cursor, count_mode="EXACT"):
    """
    count_mode:
      - EXACT -> SELECT COUNT(*)
      - FAST  -> USER_TABLES.NUM_ROWS
    """
    count_mode = count_mode.upper().strip()
    if count_mode not in ("EXACT", "FAST"):
        raise ValueError("count_mode harus EXACT atau FAST")

    tables = get_all_tables(cursor)
    stats = get_table_stats(cursor)

    results = []
    total_rows = 0

    log("\n" + "=" * 100)
    log(f"START TABLE AUDIT | MODE = {count_mode}")
    log("=" * 100)

    for i, table_name in enumerate(tables, start=1):
        try:
            log(f"[{i}/{len(tables)}] Checking table: {table_name}")

            if count_mode == "FAST":
                row_count = stats.get(table_name, {}).get("num_rows")
                if row_count is None:
                    row_count = 0
            else:
                row_count = count_table_rows_exact(cursor, table_name)

            total_rows += (row_count or 0)

            results.append({
                "table_name": table_name,
                "row_count": int(row_count or 0),
                "last_analyzed": stats.get(table_name, {}).get("last_analyzed"),
                "count_mode": count_mode,
                "error": None
            })

        except Exception as e:
            log(f"❌ Gagal count table {table_name}: {str(e)}")
            results.append({
                "table_name": table_name,
                "row_count": None,
                "last_analyzed": stats.get(table_name, {}).get("last_analyzed"),
                "count_mode": count_mode,
                "error": str(e)
            })

    return {
        "schema": cursor.connection.username.upper(),
        "table_count": len(tables),
        "total_rows": total_rows,
        "count_mode": count_mode,
        "audit_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "tables": results
    }

# =========================================================
# EXPORT CSV
# =========================================================
def export_csv(audit_result: dict, output_dir: Path):
    output_dir.mkdir(parents=True, exist_ok=True)
    file_path = output_dir / f"table_audit_{audit_result['schema']}_{now_str()}.csv"

    with open(file_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=["table_name", "row_count", "last_analyzed", "count_mode", "error"]
        )
        writer.writeheader()
        for row in audit_result["tables"]:
            writer.writerow({
                "table_name": row.get("table_name"),
                "row_count": row.get("row_count"),
                "last_analyzed": row.get("last_analyzed"),
                "count_mode": row.get("count_mode"),
                "error": row.get("error")
            })

    return file_path

# =========================================================
# EXPORT JSON
# =========================================================
def export_json(audit_result: dict, output_dir: Path):
    output_dir.mkdir(parents=True, exist_ok=True)
    file_path = output_dir / f"table_audit_{audit_result['schema']}_{now_str()}.json"

    with open(file_path, "w", encoding="utf-8") as f:
        json.dump(audit_result, f, indent=2, ensure_ascii=False)

    return file_path

# =========================================================
# EXCEL HELPERS
# =========================================================
def auto_fit_columns(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            try:
                val = "" if cell.value is None else str(cell.value)
                if len(val) > max_length:
                    max_length = len(val)
            except:
                pass

        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

def style_header(ws, row_num=1):
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    align = Alignment(horizontal="center", vertical="center")

    for cell in ws[row_num]:
        cell.fill = fill
        cell.font = font
        cell.alignment = align

def add_auto_filter(ws):
    ws.auto_filter.ref = ws.dimensions

# =========================================================
# EXPORT EXCEL
# =========================================================
def export_excel(audit_result: dict, output_dir: Path):
    output_dir.mkdir(parents=True, exist_ok=True)
    file_path = output_dir / f"table_audit_{audit_result['schema']}_{now_str()}.xlsx"

    wb = Workbook()

    # Remove default sheet
    default_ws = wb.active
    wb.remove(default_ws)

    # -----------------------------------------------------
    # SHEET 1: SUMMARY
    # -----------------------------------------------------
    ws_summary = wb.create_sheet("SUMMARY")

    tables = audit_result["tables"]
    success_rows = [t for t in tables if t.get("row_count") is not None]
    failed_rows = [t for t in tables if t.get("row_count") is None]
    empty_tables = [t for t in success_rows if t["row_count"] == 0]
    non_empty_tables = [t for t in success_rows if t["row_count"] > 0]
    biggest_table = max(non_empty_tables, key=lambda x: x["row_count"], default=None)

    summary_data = [
        ["SCHEMA", audit_result["schema"]],
        ["AUDIT TIME", audit_result["audit_time"]],
        ["COUNT MODE", audit_result["count_mode"]],
        ["TOTAL TABLES", audit_result["table_count"]],
        ["TOTAL ROWS", audit_result["total_rows"]],
        ["TABLES WITH DATA", len(non_empty_tables)],
        ["EMPTY TABLES", len(empty_tables)],
        ["FAILED TABLE COUNT", len(failed_rows)],
        ["BIGGEST TABLE", biggest_table["table_name"] if biggest_table else None],
        ["BIGGEST TABLE ROWS", biggest_table["row_count"] if biggest_table else None],
    ]

    ws_summary.append(["METRIC", "VALUE"])
    for row in summary_data:
        ws_summary.append(row)

    style_header(ws_summary, 1)
    ws_summary.freeze_panes = "A2"
    add_auto_filter(ws_summary)
    auto_fit_columns(ws_summary)

    # -----------------------------------------------------
    # SHEET 2: TABLE DETAILS
    # -----------------------------------------------------
    ws_detail = wb.create_sheet("TABLE_DETAILS")

    headers = ["table_name", "row_count", "last_analyzed", "count_mode", "error"]
    ws_detail.append(headers)

    for row in sorted(tables, key=lambda x: (x["row_count"] is None, -(x["row_count"] or 0), x["table_name"])):
        ws_detail.append([
            row.get("table_name"),
            row.get("row_count"),
            row.get("last_analyzed"),
            row.get("count_mode"),
            row.get("error")
        ])

    style_header(ws_detail, 1)
    ws_detail.freeze_panes = "A2"
    add_auto_filter(ws_detail)
    auto_fit_columns(ws_detail)

    # Format number column
    for row in ws_detail.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = '#,##0'

    # -----------------------------------------------------
    # SHEET 3: EMPTY TABLES
    # -----------------------------------------------------
    ws_empty = wb.create_sheet("EMPTY_TABLES")
    ws_empty.append(headers)

    for row in sorted(empty_tables, key=lambda x: x["table_name"]):
        ws_empty.append([
            row.get("table_name"),
            row.get("row_count"),
            row.get("last_analyzed"),
            row.get("count_mode"),
            row.get("error")
        ])

    style_header(ws_empty, 1)
    ws_empty.freeze_panes = "A2"
    add_auto_filter(ws_empty)
    auto_fit_columns(ws_empty)

    # -----------------------------------------------------
    # SHEET 4: FAILED TABLES
    # -----------------------------------------------------
    ws_failed = wb.create_sheet("FAILED_TABLES")
    ws_failed.append(headers)

    for row in sorted(failed_rows, key=lambda x: x["table_name"]):
        ws_failed.append([
            row.get("table_name"),
            row.get("row_count"),
            row.get("last_analyzed"),
            row.get("count_mode"),
            row.get("error")
        ])

    style_header(ws_failed, 1)
    ws_failed.freeze_panes = "A2"
    add_auto_filter(ws_failed)
    auto_fit_columns(ws_failed)

    wb.save(file_path)
    return file_path

# =========================================================
# REPORT
# =========================================================
def print_summary(audit_result: dict):
    tables = audit_result["tables"]

    success_rows = [t for t in tables if t.get("row_count") is not None]
    failed_rows = [t for t in tables if t.get("row_count") is None]

    empty_tables = [t for t in success_rows if t["row_count"] == 0]
    non_empty_tables = [t for t in success_rows if t["row_count"] > 0]

    biggest_table = max(non_empty_tables, key=lambda x: x["row_count"], default=None)

    print("\n" + "=" * 100)
    print("ORACLE TABLE AUDIT SUMMARY")
    print("=" * 100)
    print(f"SCHEMA              : {audit_result['schema']}")
    print(f"COUNT MODE          : {audit_result['count_mode']}")
    print(f"TOTAL TABLES        : {audit_result['table_count']}")
    print(f"TOTAL ROWS          : {audit_result['total_rows']:,}")
    print(f"TABLES WITH DATA    : {len(non_empty_tables)}")
    print(f"EMPTY TABLES        : {len(empty_tables)}")
    print(f"FAILED TABLE COUNT  : {len(failed_rows)}")

    if biggest_table:
        print(f"BIGGEST TABLE       : {biggest_table['table_name']} ({biggest_table['row_count']:,} rows)")

    print("\nTOP 20 BIGGEST TABLES")
    print("-" * 100)
    sorted_tables = sorted(
        success_rows,
        key=lambda x: x["row_count"],
        reverse=True
    )

    for row in sorted_tables[:20]:
        print(f"{row['table_name']:<40} {row['row_count']:>20,}")

    if failed_rows:
        print("\nFAILED TABLES")
        print("-" * 100)
        for row in failed_rows:
            print(f"{row['table_name']:<40} ERROR: {row.get('error')}")

# =========================================================
# MAIN
# =========================================================
def main():
    # Ganti mode di sini:
    COUNT_MODE = "EXACT"   # pilihan: EXACT / FAST

    cfg = load_config()

    print("=" * 100)
    print("ORACLE TABLE AUDIT TOOL")
    print("=" * 100)
    print(f"TARGET SCHEMA : {cfg['user']}")
    print(f"DSN           : {cfg['dsn']}")
    print(f"COUNT MODE    : {COUNT_MODE}")
    print("=" * 100)

    conn = connect_oracle(cfg)
    cursor = conn.cursor()

    try:
        audit_result = audit_tables(cursor, count_mode=COUNT_MODE)
        print_summary(audit_result)

        csv_path = export_csv(audit_result, OUTPUT_DIR)
        json_path = export_json(audit_result, OUTPUT_DIR)
        excel_path = export_excel(audit_result, OUTPUT_DIR)

        print("\n" + "=" * 100)
        print("EXPORT FILES")
        print("=" * 100)
        print(f"CSV   : {csv_path}")
        print(f"JSON  : {json_path}")
        print(f"EXCEL : {excel_path}")

    except Exception as e:
        print("\n🔥 FATAL ERROR")
        print(str(e))
        print(traceback.format_exc())

    finally:
        cursor.close()
        conn.close()
        print("\n🔒 Connection closed.")

if __name__ == "__main__":
    main()